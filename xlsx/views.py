from rest_framework import viewsets
from rest_framework.views import APIView
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from reverseGeoCodeing.settings import BASE_DIR, STATIC_ROOT
from utils.line_converter import convertLatLngLine


class ConvertXlsxView(APIView):
    def post(self,request, format=None):
            file=request.FILES['file']
            
            import openpyxl
         
            wb= openpyxl.load_workbook(file)
            
            sheetNames=wb.sheetnames

            for name in sheetNames:

                sheet=wb[name]
                for i in range(1,sheet.max_row):
                    try:
                        value=sheet.cell(row=i,column=1).value
                        if value:
                            newLine=convertLatLngLine(str(value))
                            sheet.cell(row=i,column=1).value=newLine

                        value=sheet.cell(row=i,column=2).value

                        if value:
                            newLine=convertLatLngLine(str(value))
                            sheet.cell(row=i,column=2).value=newLine

                      
                    except:
                        pass

            import uuid
            
            randomName = str(uuid.uuid4())
            xslxfilename="static/"+randomName+".xlsx"
            xslxpath=request.get_host()+"/"+xslxfilename
            pdffilename="static/"+randomName+".pdf"
            pdfpath=request.get_host()+"/"+pdffilename
            wb.save(xslxfilename)
            
     
            return Response({"xlsx":xslxpath,"pdf":pdfpath})
            # return(Response(str(file)))








# def main():
#     import openpyxl
#     wb= openpyxl.load_workbook('rapport.xlsx')

    

#     sheetNames=wb.sheetnames

#     for name in sheetNames:
#         sheet=wb[name]
#         for i in range(1,50):
#             try:
#                 newLine=extractLine(str(sheet.cell(row=i,column=1).value))
#                 sheet.cell(row=i,column=1).value=newLine
#                 newLine=extractLine(str(sheet.cell(row=i,column=2).value))
#                 sheet.cell(row=i,column=2).value=newLine
#             except:
#                 pass
#     wb.save("new.xlsx")